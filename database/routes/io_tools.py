"""
Generic Import / Export for master pages.

One blueprint (`io_bp`) exposes:
    GET  /io/<key>/export?fmt=xlsx|csv     -> download all rows of a registered model
    POST /io/<key>/import                  -> import rows (add-new-only by default)
    GET  /io/<key>/template?fmt=xlsx|csv   -> download an empty template

Each master page registers itself once via `register_io(...)`, declaring its
model, which columns to include, and the unique key used to skip existing rows
on import. No per-page route code is needed beyond that one registration.

Purchases and Sales documents are intentionally NOT registered here.
"""

import io as _io
import csv as _csv
from datetime import date, datetime

from flask import Blueprint, request, jsonify, send_file, session
from flask_login import login_required

from models import db

io_bp = Blueprint('io', __name__, url_prefix='/io')

# key -> spec dict
_REGISTRY = {}


def register_io(key, model, columns, unique='code', label=None, parents=None,
                coerce=None):
    """Register a model for generic import/export.

    key      : short url-safe id, e.g. 'work-allocations'
    model    : the SQLAlchemy model class
    columns  : list of (header, attribute) pairs to export/import
    unique   : attribute used to detect existing rows on import (skip if present)
    label    : human title (defaults to key)
    parents  : optional list of (header, fk_attr, parent_model, parent_key_attr,
               parent_id_attr) to resolve a parent by a natural key on import
    coerce   : optional dict {attr: callable(value)->value} for type conversion
    """
    _REGISTRY[key] = {
        'model': model, 'columns': columns, 'unique': unique,
        'label': label or key, 'parents': parents or [],
        'coerce': coerce or {},
    }


def _t(en, ar):
    return ar if session.get('lang') == 'ar' else en


def _val_out(v):
    if v is None:
        return ''
    if isinstance(v, (date, datetime)):
        return v.isoformat()
    if isinstance(v, bool):
        return '1' if v else '0'
    return v


def _coerce_value(attr, raw, spec):
    fn = spec['coerce'].get(attr)
    if fn:
        try:
            return fn(raw)
        except Exception:
            return None
    return (str(raw).strip() if raw is not None else None) or None


@io_bp.route('/<key>/export')
@login_required
def io_export(key):
    spec = _REGISTRY.get(key)
    if not spec:
        return jsonify({'ok': False, 'error': 'Unknown export'}), 404
    fmt = (request.args.get('fmt') or 'xlsx').lower()
    model = spec['model']
    rows = model.query.all()
    headers = [h for h, _ in spec['columns']]

    # For child exports: build a parent-id -> parent-code map so the
    # '_parent' placeholder column can show the parent's natural key.
    parent_code_map = {}
    parent_fk_attr = None
    for phdr, fk_attr, pmodel, pkey_attr, pid_attr in spec['parents']:
        parent_fk_attr = fk_attr if fk_attr == '_parent' else pid_attr
        # map the child's FK id -> parent code
        parent_code_map = {getattr(p, 'id'): getattr(p, pkey_attr, '')
                           for p in pmodel.query.with_entities(pmodel.id, getattr(pmodel, pkey_attr)).all()}
        _child_fk_id_attr = pid_attr  # actual FK column on the child (e.g. seller_id)

    def cell(r, attr):
        if attr == '_parent' and spec['parents']:
            fk_id = getattr(r, spec['parents'][0][4], None)  # pid_attr = child FK col
            return parent_code_map.get(fk_id, '')
        return getattr(r, attr, '')

    if fmt == 'csv':
        buf = _io.StringIO()
        w = _csv.writer(buf)
        w.writerow(headers)
        for r in rows:
            w.writerow([_val_out(cell(r, a)) for _, a in spec['columns']])
        data = buf.getvalue().encode('utf-8-sig')
        return send_file(_io.BytesIO(data), as_attachment=True,
                         download_name=f'{key}.csv', mimetype='text/csv')

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = spec['label'][:31] or key[:31]
    fill = PatternFill('solid', fgColor='1E3A5F')
    font = Font(color='FFFFFF', bold=True, size=10)
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = fill; c.font = font
        ws.column_dimensions[get_column_letter(i)].width = max(14, len(h) + 3)
    ws.freeze_panes = 'A2'
    for r_i, r in enumerate(rows, 2):
        for c_i, (_, a) in enumerate(spec['columns'], 1):
            ws.cell(row=r_i, column=c_i, value=_val_out(cell(r, a)))
    out = _io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out, as_attachment=True, download_name=f'{key}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def _sample_value(header):
    """A plausible example value for a column, shown as a guide row in templates."""
    h = (header or '').lower()
    if 'code' in h and 'account' in h:
        return 'A2-01-03-003'
    if h.endswith('code') or h == 'code':
        return 'ABC-0001'
    if 'name (ar)' in h or h.endswith('(ar)'):
        return 'مثال'
    if 'name' in h:
        return 'Example Name'
    if 'email' in h:
        return 'name@example.com'
    if 'phone' in h or 'fax' in h or 'mobile' in h:
        return '0500000000'
    if 'vat' in h:
        return '300000000000003'
    if 'crn' in h:
        return '1010000000'
    if 'iban' in h:
        return 'SA0000000000000000000000'
    if 'swift' in h:
        return 'ABCDSARIXXX'
    if 'account number' in h:
        return '1234567890'
    if 'date' in h:
        return '2026-01-31'
    if 'rate' in h or 'salary' in h or 'multiply' in h or 'mrp' in h:
        return '0'
    if 'active' in h or 'primary' in h:
        return '1'
    if 'status' in h:
        return 'active'
    if 'year' in h:
        return '2026'
    if 'city' in h:
        return 'Riyadh'
    if 'country' in h:
        return 'Saudi Arabia'
    if 'website' in h:
        return 'https://example.com'
    return 'Sample'


_SAMPLE_SENTINELS = {'ABC-0001', 'A2-01-03-003', 'Example Name',
                     '↑ Sample row — delete before importing'}


def _is_note_row(row):
    """True if a row is the template's sample/note line (should not import)."""
    for v in row.values():
        if v is None:
            continue
        s = str(v).strip()
        if s.startswith('↑') or s in _SAMPLE_SENTINELS:
            return True
    return False


@io_bp.route('/<key>/template')
@login_required
def io_template(key):
    """Empty file with just the headers, for filling in and importing."""
    spec = _REGISTRY.get(key)
    if not spec:
        return jsonify({'ok': False, 'error': 'Unknown template'}), 404
    fmt = (request.args.get('fmt') or 'xlsx').lower()
    headers = [h for h, _ in spec['columns']]
    sample = [_sample_value(h) for h in headers]
    if fmt == 'csv':
        buf = _io.StringIO(); w = _csv.writer(buf)
        w.writerow(headers)
        w.writerow(sample)
        data = buf.getvalue().encode('utf-8-sig')
        return send_file(_io.BytesIO(data), as_attachment=True,
                         download_name=f'{key}_template.csv', mimetype='text/csv')
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    wb = Workbook(); ws = wb.active; ws.title = 'Template'
    fill = PatternFill('solid', fgColor='1E3A5F'); font = Font(color='FFFFFF', bold=True)
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h); c.fill = fill; c.font = font
        ws.column_dimensions[get_column_letter(i)].width = max(14, len(h) + 3)
    sample_font = Font(italic=True, color='94A3B8')
    for i, v in enumerate(sample, 1):
        ws.cell(row=2, column=i, value=v).font = sample_font
    ws.cell(row=3, column=1, value='↑ Sample row — delete before importing').font = Font(italic=True, color='DC2626')
    out = _io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out, as_attachment=True, download_name=f'{key}_template.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def _read_uploaded(f):
    """Return (headers, list-of-row-dicts) from an uploaded csv/xlsx."""
    name = f.filename.lower()
    if name.endswith('.csv'):
        text = f.read().decode('utf-8-sig', errors='replace')
        reader = _csv.DictReader(_io.StringIO(text))
        return list(reader)
    if name.endswith('.xlsx') or name.endswith('.xlsm'):
        from openpyxl import load_workbook
        wb = load_workbook(f, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        try:
            headers = [str(h).strip() if h is not None else '' for h in next(it)]
        except StopIteration:
            return []
        rows = []
        for raw in it:
            if raw is None:
                continue
            rows.append({headers[i]: raw[i] for i in range(min(len(headers), len(raw)))})
        return rows
    raise ValueError('Unsupported file type. Use .csv or .xlsx')


@io_bp.route('/<key>/import', methods=['POST'])
@login_required
def io_import(key):
    spec = _REGISTRY.get(key)
    if not spec:
        return jsonify({'ok': False, 'error': 'Unknown import'}), 404
    f = request.files.get('file')
    if not f or not f.filename:
        return jsonify({'ok': False, 'error': 'No file uploaded.'}), 400

    model = spec['model']
    ukey = spec['unique']
    hdr_by_attr = {a: h for h, a in spec['columns']}

    try:
        rows = _read_uploaded(f)
    except ValueError as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

    # Existing unique values to skip.
    existing = set()
    if ukey and hasattr(model, ukey):
        existing = {getattr(r, ukey) for r in
                    model.query.with_entities(getattr(model, ukey)).all()}

    # Parent lookup maps.
    parent_maps = {}
    for phdr, fk_attr, pmodel, pkey_attr, pid_attr in spec['parents']:
        parent_maps[phdr] = (fk_attr, pid_attr,
                             {getattr(p, pkey_attr): p.id for p in
                              pmodel.query.with_entities(getattr(pmodel, pkey_attr), pmodel.id).all()})

    added = skipped = 0
    errors = []
    for idx, row in enumerate(rows, 2):
        if _is_note_row(row):
            continue
        uheader = hdr_by_attr.get(ukey)
        uval = (str(row.get(uheader, '')).strip() if uheader else '')
        if ukey and not uval:
            continue  # skip blank key rows
        if ukey and uval in existing:
            skipped += 1
            continue

        obj = model()
        for header, attr in spec['columns']:
            if attr in (fk for _, fk, *_ in spec['parents']):
                continue  # parents handled below
            if header not in row:
                continue
            setattr(obj, attr, _coerce_value(attr, row.get(header), spec))

        ok = True
        for phdr, (fk_attr, pid_attr, pmap) in parent_maps.items():
            pval = str(row.get(phdr, '')).strip()
            if pval and pval in pmap:
                if fk_attr != '_parent':
                    setattr(obj, fk_attr, pval)
                setattr(obj, pid_attr, pmap[pval])
            elif pval:
                errors.append(f'Row {idx}: parent "{pval}" not found; skipped.')
                ok = False
                break
        if not ok:
            continue

        db.session.add(obj)
        if ukey:
            existing.add(uval)
        added += 1

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 500

    return jsonify({'ok': True, 'added': added, 'skipped': skipped,
                    'errors': errors[:50]})


# ══════════════════════════════════════════════════════════════════
#  BUNDLES: parent + child tables as ONE multi-sheet workbook
# ══════════════════════════════════════════════════════════════════
_BUNDLES = {}


def register_bundle(key, parent_model, parent_unique, parent_columns,
                    children, label=None, parent_coerce=None):
    """Register a parent-with-children bundle for multi-sheet import/export.

    key            : url id, e.g. 'seller-full'
    parent_model   : the parent model (e.g. Seller)
    parent_unique  : parent's natural key attribute (e.g. 'seller_code')
    parent_columns : list of (header, attr) for the parent sheet
    children       : list of dicts, each:
        {
          'sheet'       : sheet name (e.g. 'Banks'),
          'model'       : child model,
          'fk'          : child FK attr to parent id (e.g. 'seller_id'),
          'parent_ref'  : header used in the child sheet to name the parent
                          (e.g. 'Seller Code'),
          'columns'     : list of (header, attr) for the child's own fields,
          'coerce'      : optional {attr: fn},
        }
    label          : workbook title
    """
    _BUNDLES[key] = {
        'parent_model': parent_model, 'parent_unique': parent_unique,
        'parent_columns': parent_columns, 'children': children,
        'label': label or key, 'parent_coerce': parent_coerce or {},
    }


@io_bp.route('/bundle/<key>/export')
@login_required
def bundle_export(key):
    spec = _BUNDLES.get(key)
    if not spec:
        return jsonify({'ok': False, 'error': 'Unknown bundle'}), 404

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    fill = PatternFill('solid', fgColor='1E3A5F')
    font = Font(color='FFFFFF', bold=True, size=10)

    def write_sheet(ws, headers, rows_values, note=None):
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.fill = fill; c.font = font
            ws.column_dimensions[get_column_letter(i)].width = max(14, len(h) + 3)
        ws.freeze_panes = 'A2'
        for r_i, vals in enumerate(rows_values, 2):
            for c_i, v in enumerate(vals, 1):
                ws.cell(row=r_i, column=c_i, value=_val_out(v))
        if note:
            ws.cell(row=max(2, len(rows_values) + 3), column=1, value=note)

    def safe_rows(model, attrs):
        """Query only the given attributes; return list of value-tuples.
        Falls back to per-object access; never raises."""
        try:
            cols = [getattr(model, a) for a in attrs]
            return [tuple(rv) for rv in db.session.query(*cols).all()], None
        except Exception as e:
            try:
                out = []
                for r in model.query.all():
                    out.append(tuple(getattr(r, a, '') for a in attrs))
                return out, None
            except Exception as e2:
                return [], str(e2 or e)

    wb = Workbook()

    # ── Parent sheet ──
    pmodel = spec['parent_model']
    pcols = spec['parent_columns']
    punique = spec['parent_unique']
    ws = wb.active
    ws.title = (spec['label'][:28] or key[:28])
    # include id first so we can map children, then the declared columns
    p_attrs = ['id', punique] + [a for _, a in pcols]
    p_rows, p_err = safe_rows(pmodel, p_attrs)
    id_to_key = {row[0]: row[1] for row in p_rows}
    write_sheet(ws, [h for h, _ in pcols],
                [list(row[2:]) for row in p_rows],
                note=(f'Could not fully read parent: {p_err}' if p_err else None))

    # ── Child sheets (always created) ──
    for child in spec['children']:
        cws = wb.create_sheet(title=child['sheet'][:31])
        if child.get('standalone'):
            # Not linked to the parent: export all rows of its own columns.
            headers = [h for h, _ in child['columns']]
            c_attrs = [a for _, a in child['columns']]
            c_rows, c_err = safe_rows(child['model'], c_attrs)
            values = [list(row) for row in c_rows]
            write_sheet(cws, headers, values,
                        note=(f'Could not read {child["sheet"]}: {c_err}' if c_err else None))
            continue
        headers = [child['parent_ref']] + [h for h, _ in child['columns']]
        c_attrs = [child['fk']] + [a for _, a in child['columns']]
        c_rows, c_err = safe_rows(child['model'], c_attrs)
        values = [[id_to_key.get(row[0], '')] + list(row[1:]) for row in c_rows]
        write_sheet(cws, headers, values,
                    note=(f'Could not read {child["sheet"]}: {c_err}' if c_err else None))

    out = _io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out, as_attachment=True, download_name=f'{key}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@io_bp.route('/bundle/<key>/template')
@login_required
def bundle_template(key):
    """Multi-sheet template: parent + child sheets, each with a sample row."""
    spec = _BUNDLES.get(key)
    if not spec:
        return jsonify({'ok': False, 'error': 'Unknown bundle'}), 404
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    fill = PatternFill('solid', fgColor='1E3A5F')
    font = Font(color='FFFFFF', bold=True)
    sfont = Font(italic=True, color='94A3B8')

    def build(ws, headers):
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h); c.fill = fill; c.font = font
            ws.column_dimensions[get_column_letter(i)].width = max(14, len(h) + 3)
        for i, h in enumerate(headers, 1):
            ws.cell(row=2, column=i, value=_sample_value(h)).font = sfont
        ws.cell(row=3, column=1, value='↑ Sample row — delete before importing').font = Font(italic=True, color='DC2626')

    wb = Workbook()
    ws = wb.active
    ws.title = (spec['label'][:28] or key[:28])
    build(ws, [h for h, _ in spec['parent_columns']])
    for child in spec['children']:
        cws = wb.create_sheet(title=child['sheet'][:31])
        if child.get('standalone'):
            build(cws, [h for h, _ in child['columns']])
        else:
            build(cws, [child['parent_ref']] + [h for h, _ in child['columns']])
    out = _io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out, as_attachment=True, download_name=f'{key}_template.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@io_bp.route('/bundle/<key>/import', methods=['POST'])
@login_required
def bundle_import(key):
    spec = _BUNDLES.get(key)
    if not spec:
        return jsonify({'ok': False, 'error': 'Unknown bundle'}), 404
    f = request.files.get('file')
    if not f or not f.filename:
        return jsonify({'ok': False, 'error': 'No file uploaded.'}), 400
    name = f.filename.lower()
    if not (name.endswith('.xlsx') or name.endswith('.xlsm')):
        return jsonify({'ok': False, 'error': 'Bundle import needs an .xlsx file.'}), 400

    from openpyxl import load_workbook
    try:
        wb = load_workbook(f, data_only=True, read_only=True)
    except Exception as e:
        return jsonify({'ok': False, 'error': f'Could not read file: {e}'}), 400

    def sheet_rows(ws):
        it = ws.iter_rows(values_only=True)
        try:
            headers = [str(h).strip() if h is not None else '' for h in next(it)]
        except StopIteration:
            return []
        out = []
        for raw in it:
            if raw is None:
                continue
            out.append({headers[i]: raw[i] for i in range(min(len(headers), len(raw)))})
        return out

    result = {'parent_added': 0, 'parent_skipped': 0, 'children': {}, 'errors': []}

    pmodel = spec['parent_model']
    punique = spec['parent_unique']
    pcols = spec['parent_columns']
    phdr_by_attr = {a: h for h, a in pcols}
    puheader = phdr_by_attr.get(punique)

    # ── Parent sheet (first sheet, by label or first) ──
    parent_sheet = wb.sheetnames[0]
    existing_parent = {getattr(p, punique) for p in
                       pmodel.query.with_entities(getattr(pmodel, punique)).all()}
    for row in sheet_rows(wb[parent_sheet]):
        if _is_note_row(row):
            continue
        uval = str(row.get(puheader, '')).strip() if puheader else ''
        if not uval:
            continue
        if uval in existing_parent:
            result['parent_skipped'] += 1
            continue
        obj = pmodel()
        pcoerce = spec.get('parent_coerce', {})
        for header, attr in pcols:
            if header in row:
                fn = pcoerce.get(attr)
                if fn:
                    try: val = fn(row[header])
                    except Exception: val = None
                else:
                    val = (str(row[header]).strip() if row[header] is not None else None) or None
                setattr(obj, attr, val)
        db.session.add(obj)
        existing_parent.add(uval)
        result['parent_added'] += 1
    db.session.flush()

    # refreshed key -> id map
    key_to_id = {getattr(p, punique): p.id for p in
                 pmodel.query.with_entities(getattr(pmodel, punique), pmodel.id).all()}

    # ── Child sheets ──
    for child in spec['children']:
        sheet = child['sheet']
        added = 0
        skipped = 0
        if sheet not in wb.sheetnames:
            result['children'][sheet] = {'added': 0, 'skipped': 0}
            continue
        cmodel = child['model']
        coerce = child.get('coerce', {})

        # Standalone sheet: independent master, not linked to the parent.
        if child.get('standalone'):
            uattr = child.get('unique')
            existing = set()
            if uattr and hasattr(cmodel, uattr):
                existing = {getattr(r, uattr) for r in
                            cmodel.query.with_entities(getattr(cmodel, uattr)).all()}
            uheader = None
            for h, a in child['columns']:
                if a == uattr:
                    uheader = h; break
            for idx, row in enumerate(sheet_rows(wb[sheet]), 2):
                if _is_note_row(row):
                    continue
                uval = str(row.get(uheader, '')).strip() if uheader else ''
                if uattr and not uval:
                    continue
                if uattr and uval in existing:
                    skipped += 1
                    continue
                obj = cmodel()
                for header, attr in child['columns']:
                    if header not in row:
                        continue
                    raw = row[header]
                    fn = coerce.get(attr)
                    if fn:
                        try: val = fn(raw)
                        except Exception: val = None
                    else:
                        val = (str(raw).strip() if raw is not None else None) or None
                    setattr(obj, attr, val)
                db.session.add(obj)
                if uattr:
                    existing.add(uval)
                added += 1
            result['children'][sheet] = {'added': added, 'skipped': skipped}
            continue

        for idx, row in enumerate(sheet_rows(wb[sheet]), 2):
            if _is_note_row(row):
                continue
            pref = str(row.get(child['parent_ref'], '')).strip()
            if not pref:
                continue
            pid = key_to_id.get(pref)
            if not pid:
                result['errors'].append(f"{sheet} row {idx}: parent '{pref}' not found; skipped.")
                skipped += 1
                continue
            obj = cmodel()
            setattr(obj, child['fk'], pid)
            for header, attr in child['columns']:
                if header not in row:
                    continue
                raw = row[header]
                fn = coerce.get(attr)
                if fn:
                    try:
                        val = fn(raw)
                    except Exception:
                        val = None
                else:
                    val = (str(raw).strip() if raw is not None else None) or None
                setattr(obj, attr, val)
            resolver = child.get('resolve')
            if resolver:
                try:
                    ok = resolver(obj, row, pid)
                    if ok is False:
                        skipped += 1
                        continue
                except Exception as e:
                    result['errors'].append(f"{sheet} row {idx}: {e}")
                    skipped += 1
                    continue
            db.session.add(obj)
            added += 1
        result['children'][sheet] = {'added': added, 'skipped': skipped}

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 500

    result['ok'] = True
    result['errors'] = result['errors'][:50]
    return jsonify(result)


# ══════════════════════════════════════════════════════════════════
#  INLINE CELL EDITING — update one field of one row
# ══════════════════════════════════════════════════════════════════
# Fields that must never be editable inline, regardless of registration.
_NEVER_EDITABLE = {'id', 'seller_code', 'buyer_code', 'vendor_code',
                   'employee_code', 'item_code', 'created_at', 'updated_at',
                   'created_by', 'je_no', 'sheet_no', 'year',
                   # computed / derived — must not be hand-edited
                   'total_salary', 'salary_payable', 'monthly_salary',
                   'ot_amount', 'total_hours',
                   'financial_year', 'range', 'status_locked'}


@io_bp.route('/<key>/update-cell', methods=['POST'])
@login_required
def io_update_cell(key):
    """Update a single column of a single row for a registered model.

    JSON body: { "id": <row id>, "field": "<attr>", "value": <new value> }
    Only attributes declared in the model's io registration are allowed,
    minus the unique key and a hard blocklist. Returns the saved value.
    """
    spec = _REGISTRY.get(key)
    if not spec:
        return jsonify({'ok': False, 'error': 'Unknown grid'}), 404

    data = request.get_json(silent=True) or {}
    row_id = data.get('id')
    field = (data.get('field') or '').strip()
    value = data.get('value')

    if row_id is None or not field:
        return jsonify({'ok': False, 'error': 'Missing id or field.'}), 400

    model = spec['model']
    # Editable = any real column on the model, minus the protected blocklist.
    # (Registered columns are a subset; we allow the full model so every
    #  genuine field can be edited inline, while _NEVER_EDITABLE stays blocked.)
    model_cols = {c.name for c in model.__table__.columns}
    editable = model_cols - _NEVER_EDITABLE
    if field not in editable:
        return jsonify({'ok': False, 'error': f'Field "{field}" is not editable.'}), 403

    obj = db.session.get(model, row_id)
    if not obj:
        return jsonify({'ok': False, 'error': 'Row not found.'}), 404

    # Coerce the value using the same rule as import, if provided.
    coerce_fn = spec.get('coerce', {}).get(field)
    try:
        if coerce_fn:
            new_val = coerce_fn(value)
        else:
            new_val = (str(value).strip() if value is not None else None)
            if new_val == '':
                new_val = None
        setattr(obj, field, new_val)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 500

    return jsonify({'ok': True, 'id': row_id, 'field': field,
                    'value': _val_out(getattr(obj, field, ''))})