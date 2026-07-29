"""Chart of Accounts module — Level 1 and Level 2.

Routes:
    /coa/level-one          list Level 1
    /coa/level-one/add      create Level 1
    /coa/level-one/<id>/edit
    /coa/level-one/<id>/delete
    /coa/level-two          list Level 2
    /coa/level-two/add      create Level 2  (code auto-generated)
    /coa/level-two/<id>/edit
    /coa/level-two/<id>/delete
    /coa/level-two/next-code?level_one_id=..   preview next auto code

Business rules enforced server-side (never trust the client):
    * Level 1 ``code_length`` is always 1; ``code`` is fixed after creation.
    * Level 2 ``code_length`` is always 2; ``description`` is always
      'Heading Account'; ``code`` is generated as <L1 code><n>, with an
      independent sequence per Level 1.
"""
import re
from functools import wraps
from flask import (Blueprint, render_template, request, redirect, url_for,
                   flash, session, jsonify, send_file)
from flask_login import login_required, current_user

from models import db, LevelOne, LevelTwo

coa_bp = Blueprint('coa', __name__, url_prefix='/coa')


# ── Chart-of-Accounts fixed drawer order (A,L,E,R,C,O,F,I) ────────
# Applied everywhere codes are sorted, in grid and tree views.
COA_ORDER = ['A', 'L', 'E', 'R', 'C', 'O', 'F', 'I']


def coa_sort_key(code_col):
    """
    SQLAlchemy ORDER BY expression that sorts by the leading Level-1 letter
    in the fixed order A,L,E,R,C,O,F,I, then by the full code naturally.
    Unknown leading letters sort last, then alphabetically.
    """
    from sqlalchemy import case, func
    lead = func.upper(func.substr(code_col, 1, 1))
    whens = {letter: idx for idx, letter in enumerate(COA_ORDER)}
    rank = case(whens, value=lead, else_=len(COA_ORDER))
    return rank, code_col


def coa_order(query, code_col):
    """Apply the fixed COA ordering to a query."""
    rank, col = coa_sort_key(code_col)
    return query.order_by(rank, col)


def coa_sort_list(rows, code_attr='code'):
    """Sort an in-memory list of ORM rows / dicts by the fixed COA order."""
    def key(r):
        code = getattr(r, code_attr, None) if not isinstance(r, dict) else r.get(code_attr)
        code = (code or '')
        lead = code[:1].upper()
        rank = COA_ORDER.index(lead) if lead in COA_ORDER else len(COA_ORDER)
        return (rank, code)
    return sorted(rows, key=key)


# ── helpers ──────────────────────────────────────────────────────
def _t(en, ar):
    return ar if session.get('lang') == 'ar' else en


def admin_required(f):
    """Only admins may create / edit / delete accounts."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_admin():
            flash(_t('Access denied.', 'الوصول مرفوض'), 'danger')
            return redirect(url_for('coa.level_one_list'))
        return f(*args, **kwargs)
    return decorated


def _status_from(req):
    """Read the Active/Inactive radio; anything unrecognised defaults to active."""
    v = (req.form.get('status') or 'active').strip().lower()
    return 'inactive' if v == 'inactive' else 'active'


def _reject_get(list_endpoint):
    """A stale template may submit an add/edit form as GET (button outside the
    <form>). Rather than show a bare 405, bounce back to the list with a hint."""
    flash(_t('That action must be submitted as a form (POST). '
             'If you see this, your page template is out of date — reload with Ctrl+F5.',
             'يجب إرسال هذا الإجراء عبر نموذج (POST). أعد تحميل الصفحة.'), 'warning')
    return redirect(url_for(list_endpoint))


def _next_level_two_code(level_one):
    """Return the next Level 2 code for a given LevelOne row.

    Sequence is independent per Level 1: reads existing Level 2 rows for that
    parent, finds the highest numeric suffix, and increments. Starts at 1 when
    none exist -> e.g. A1, A2, A3 ...
    """
    prefix = level_one.code
    rows = LevelTwo.query.filter_by(level_one_id=level_one.id).all()
    highest = 0
    for r in rows:
        # Strip the leading letter(s) and read the trailing number.
        m = re.match(r'^' + re.escape(prefix) + r'(\d+)$', r.code or '')
        if m:
            highest = max(highest, int(m.group(1)))
    candidate = f'{prefix}{highest + 1}'
    # Guard against a rare collision (e.g. manually inserted codes).
    while LevelTwo.query.filter_by(code=candidate).first():
        highest += 1
        candidate = f'{prefix}{highest + 1}'
    return candidate


# ═════════════════════════════════════════════════════════════════
#  LEVEL ONE
# ═════════════════════════════════════════════════════════════════
@coa_bp.route('/level-one')
@login_required
def level_one_list():
    pg, q_text, parent, sort, status = _paginate_filter_sort(LevelOne, LevelOne.code, request)
    return render_template('coa/level_one.html', pg=pg, rows=pg.items,
                           q=q_text, sort=sort, status=status)


@coa_bp.route('/level-one/add', methods=['GET', 'POST'])
@login_required
@admin_required
def level_one_add():
    if request.method == 'GET':
        return _reject_get('coa.level_one_list')
    from forms import LevelOneForm
    form = LevelOneForm()
    if not form.validate_on_submit():
        flash(_t('Please correct the errors and try again.',
                 'يرجى تصحيح الأخطاء والمحاولة مرة أخرى.'), 'danger')
        return redirect(url_for('coa.level_one_list'))

    code = (form.code.data or '').strip().upper()
    # Uniqueness check.
    if LevelOne.query.filter_by(code=code).first():
        flash(_t(f'Code "{code}" already exists.',
                 f'الكود "{code}" موجود بالفعل.'), 'danger')
        return redirect(url_for('coa.level_one_list'))

    drawers = (form.drawers.data or '').strip()
    if LevelOne.query.filter(db.func.lower(LevelOne.drawers) == drawers.lower()).first():
        flash(_t(f'"{drawers}" already exists.',
                 f'"{drawers}" موجود بالفعل.'), 'danger')
        return redirect(url_for('coa.level_one_list'))

    l1 = LevelOne(
        code_length=1,                       # always 1
        code=code,
        drawers=(form.drawers.data or '').strip(),
        drawers_ar=(form.drawers_ar.data or '').strip() or None,
        description=(form.description.data or '').strip(),
        description_ar=(form.description_ar.data or '').strip() or None,
        status=_status_from(request),
    )
    db.session.add(l1)
    db.session.commit()
    flash(_t(f'Level 1 account "{code}" created.',
             f'تم إنشاء حساب المستوى الأول "{code}".'), 'success')
    return redirect(url_for('coa.level_one_list'))


@coa_bp.route('/level-one/<int:id>/edit', methods=['POST'])
@login_required
@admin_required
def level_one_edit(id):
    l1 = LevelOne.query.get_or_404(id)
    from forms import LevelOneEditForm
    form = LevelOneEditForm()
    if not form.validate_on_submit():
        flash(_t('Please correct the errors and try again.',
                 'يرجى تصحيح الأخطاء والمحاولة مرة أخرى.'), 'danger')
        return redirect(url_for('coa.level_one_list'))

    # NOTE: code is fixed and cannot be changed after creation.
    l1.drawers = (form.drawers.data or '').strip()
    l1.drawers_ar = (form.drawers_ar.data or '').strip() or None
    l1.description = (form.description.data or '').strip()
    l1.description_ar = (form.description_ar.data or '').strip() or None
    l1.code_length = 1                       # keep fixed
    l1.status = _status_from(request)
    db.session.commit()
    flash(_t('Level 1 account updated.', 'تم تحديث حساب المستوى الأول.'), 'success')
    return redirect(url_for('coa.level_one_list'))


@coa_bp.route('/level-one/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def level_one_delete(id):
    l1 = LevelOne.query.get_or_404(id)
    if l1.level_twos:
        flash(_t('Cannot delete: this Level 1 has Level 2 accounts.',
                 'لا يمكن الحذف: هذا الحساب يحتوي على حسابات مستوى ثاني.'), 'danger')
        return redirect(url_for('coa.level_one_list'))
    db.session.delete(l1)
    db.session.commit()
    flash(_t('Level 1 account deleted.', 'تم حذف حساب المستوى الأول.'), 'success')
    return redirect(url_for('coa.level_one_list'))


@coa_bp.route('/level-one/data')
@login_required
def level_one_data():
    return jsonify([r.to_dict() for r in LevelOne.query.order_by(LevelOne.id).all()])


# ═════════════════════════════════════════════════════════════════
#  LEVEL TWO
# ═════════════════════════════════════════════════════════════════
@coa_bp.route('/level-two')
@login_required
def level_two_list():
    pg, q_text, parent, sort, status = _paginate_filter_sort(LevelTwo, LevelTwo.level_one_code, request)
    level_ones = coa_order(LevelOne.query, LevelOne.code).all()
    return render_template('coa/level_two.html', pg=pg, rows=pg.items,
                           level_ones=level_ones, parents=level_ones,
                           q=q_text, parent=parent, sort=sort, status=status)


@coa_bp.route('/level-two/next-code')
@login_required
def level_two_next_code():
    """Preview the auto code for the selected Level 1 (used by the form UI)."""
    l1_id = request.args.get('level_one_id', type=int)
    l1 = LevelOne.query.get(l1_id) if l1_id else None
    if not l1:
        return jsonify({'code': ''})
    return jsonify({'code': _next_level_two_code(l1), 'level_one_code': l1.code})


@coa_bp.route('/level-two/add', methods=['GET', 'POST'])
@login_required
@admin_required
def level_two_add():
    if request.method == 'GET':
        return _reject_get('coa.level_two_list')
    from forms import LevelTwoForm
    form = LevelTwoForm()
    # Populate the select choices before validating.
    form.level_one_id.choices = [
        (l1.id, f'{l1.code} — {l1.drawers}')
        for l1 in coa_order(LevelOne.query, LevelOne.code).all()
    ]
    if not form.validate_on_submit():
        flash(_t('Please correct the errors and try again.',
                 'يرجى تصحيح الأخطاء والمحاولة مرة أخرى.'), 'danger')
        return redirect(url_for('coa.level_two_list'))

    l1 = LevelOne.query.get(form.level_one_id.data)
    if not l1:
        flash(_t('Selected Level 1 account does not exist.',
                 'حساب المستوى الأول المحدد غير موجود.'), 'danger')
        return redirect(url_for('coa.level_two_list'))

    drawers = (form.drawers.data or '').strip()
    # Recommended: prevent duplicate drawers under the same Level 1.
    dup = LevelTwo.query.filter(
        LevelTwo.level_one_id == l1.id,
        db.func.lower(LevelTwo.drawers) == drawers.lower()
    ).first()
    if dup:
        flash(_t(f'"{drawers}" already exists under {l1.code}.',
                 f'"{drawers}" موجود بالفعل ضمن {l1.code}.'), 'danger')
        return redirect(url_for('coa.level_two_list'))

    # Auto-generate the code (independent sequence per Level 1).
    code = _next_level_two_code(l1)

    l2 = LevelTwo(
        code_length=2,                       # always 2
        level_one_id=l1.id,
        level_one_code=l1.code,
        code=code,
        drawers=drawers,
        drawers_ar=(form.drawers_ar.data or '').strip() or None,
        description='Heading Account',       # always
        description_ar=(form.description_ar.data or '').strip() or None,
        status=_status_from(request),
    )
    db.session.add(l2)
    db.session.commit()
    flash(_t(f'Level 2 account "{code}" created.',
             f'تم إنشاء حساب المستوى الثاني "{code}".'), 'success')
    return redirect(url_for('coa.level_two_list'))


@coa_bp.route('/level-two/<int:id>/edit', methods=['POST'])
@login_required
@admin_required
def level_two_edit(id):
    l2 = LevelTwo.query.get_or_404(id)
    from forms import LevelTwoEditForm
    form = LevelTwoEditForm()
    if not form.validate_on_submit():
        flash(_t('Please correct the errors and try again.',
                 'يرجى تصحيح الأخطاء والمحاولة مرة أخرى.'), 'danger')
        return redirect(url_for('coa.level_two_list'))

    drawers = (form.drawers.data or '').strip()
    # NOTE: code, level_one, code_length and description are all fixed.
    dup = LevelTwo.query.filter(
        LevelTwo.level_one_id == l2.level_one_id,
        db.func.lower(LevelTwo.drawers) == drawers.lower(),
        LevelTwo.id != l2.id
    ).first()
    if dup:
        flash(_t(f'"{drawers}" already exists under {l2.level_one_code}.',
                 f'"{drawers}" موجود بالفعل ضمن {l2.level_one_code}.'), 'danger')
        return redirect(url_for('coa.level_two_list'))

    l2.drawers = drawers
    l2.drawers_ar = (form.drawers_ar.data or '').strip() or None
    l2.description_ar = (form.description_ar.data or '').strip() or None
    l2.code_length = 2                       # keep fixed
    l2.description = 'Heading Account'       # keep fixed
    l2.status = _status_from(request)
    db.session.commit()
    flash(_t('Level 2 account updated.', 'تم تحديث حساب المستوى الثاني.'), 'success')
    return redirect(url_for('coa.level_two_list'))


@coa_bp.route('/level-two/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def level_two_delete(id):
    l2 = LevelTwo.query.get_or_404(id)
    db.session.delete(l2)
    db.session.commit()
    flash(_t('Level 2 account deleted.', 'تم حذف حساب المستوى الثاني.'), 'success')
    return redirect(url_for('coa.level_two_list'))


@coa_bp.route('/level-two/data')
@login_required
def level_two_data():
    return jsonify([r.to_dict() for r in
                    coa_order(LevelTwo.query, LevelTwo.code).all()])


# ═════════════════════════════════════════════════════════════════
#  LEVELS 3, 4, 5
#  Shared helpers + CRUD. All three levels behave identically:
#    * code is auto-generated (read-only)
#    * code_length and description are fixed server-side
#    * a parent cannot be deleted while it has children
#    * duplicate drawers under the same parent are rejected
# ═════════════════════════════════════════════════════════════════
from models import (LevelThree, LevelFour, LevelFive,
                    next_level_three_code, next_level_four_code, next_level_five_code)

PER_PAGE = 25


def _paginate_filter_sort(model, parent_code_col, request):
    """Apply search / parent filter / status filter / sorting / pagination.

    Query-string params:
        q        free-text search over code + drawers
        parent   exact parent-code filter
        status   active | inactive   (blank = all)
        sort     one of: code, drawers, created_at   (prefix '-' for desc)
        page     1-based page number
    """
    q_text  = (request.args.get('q') or '').strip()
    parent  = (request.args.get('parent') or '').strip()
    status  = (request.args.get('status') or '').strip().lower()
    sort    = (request.args.get('sort') or 'code').strip()
    page    = request.args.get('page', 1, type=int)

    query = model.query
    if q_text:
        like = f'%{q_text}%'
        query = query.filter(db.or_(model.code.ilike(like), model.drawers.ilike(like)))
    if parent:
        query = query.filter(parent_code_col == parent)
    if status in ('active', 'inactive'):
        query = query.filter(model.status == status)

    desc = sort.startswith('-')
    field = sort[1:] if desc else sort
    if field == 'code':
        # Fixed drawer order A,L,E,R,C,O,F,I
        rank, col = coa_sort_key(model.code)
        if desc:
            query = query.order_by(rank.desc(), col.desc())
        else:
            query = query.order_by(rank, col)
    else:
        col = {'drawers': model.drawers,
               'created_at': model.created_at}.get(field, model.code)
        query = query.order_by(col.desc() if desc else col.asc())

    return query.paginate(page=page, per_page=PER_PAGE, error_out=False), q_text, parent, sort, status


# ─── LEVEL THREE ─────────────────────────────────────────────────
@coa_bp.route('/level-three')
@login_required
def level_three_list():
    pg, q_text, parent, sort, status = _paginate_filter_sort(LevelThree, LevelThree.level_two_code, request)
    parents = coa_order(LevelTwo.query, LevelTwo.code).all()
    level_ones = coa_order(LevelOne.query, LevelOne.code).all()
    return render_template('coa/level_three.html', pg=pg, rows=pg.items,
                           parents=parents, level_ones=level_ones,
                           q=q_text, parent=parent, sort=sort, status=status)


@coa_bp.route('/level-three/next-code')
@login_required
def level_three_next_code():
    p = LevelTwo.query.get(request.args.get('level_two_id', type=int))
    return jsonify({'code': next_level_three_code(p) if p else ''})


@coa_bp.route('/level-three/add', methods=['GET', 'POST'])
@login_required
@admin_required
def level_three_add():
    if request.method == 'GET':
        return _reject_get('coa.level_three_list')
    from forms import LevelThreeForm
    form = LevelThreeForm()
    form.level_two_id.choices = [(r.id, f'{r.code} — {r.drawers}') for r in coa_order(LevelTwo.query, LevelTwo.code)]
    if not form.validate_on_submit():
        flash(_t('Please correct the errors.', 'يرجى تصحيح الأخطاء.'), 'danger')
        return redirect(url_for('coa.level_three_list'))
    parent = LevelTwo.query.get(form.level_two_id.data)
    if not parent:
        flash(_t('Parent not found.', 'الحساب الأب غير موجود.'), 'danger')
        return redirect(url_for('coa.level_three_list'))
    drawers = form.drawers.data.strip()
    if LevelThree.query.filter(LevelThree.level_two_id == parent.id,
                               db.func.lower(LevelThree.drawers) == drawers.lower()).first():
        flash(_t(f'"{drawers}" already exists under {parent.code}.',
                 f'"{drawers}" موجود بالفعل ضمن {parent.code}.'), 'danger')
        return redirect(url_for('coa.level_three_list'))
    db.session.add(LevelThree(code_length=5, level_two_id=parent.id, level_two_code=parent.code,
                              code=next_level_three_code(parent), drawers=drawers,
                              drawers_ar=(form.drawers_ar.data or '').strip() or None,
                              description='Heading Account',
                              description_ar=(form.description_ar.data or '').strip() or None,
                              status=_status_from(request)))
    db.session.commit()
    flash(_t('Level 3 account created.', 'تم إنشاء حساب المستوى الثالث.'), 'success')
    return redirect(url_for('coa.level_three_list'))


@coa_bp.route('/level-three/<int:id>/edit', methods=['POST'])
@login_required
@admin_required
def level_three_edit(id):
    row = LevelThree.query.get_or_404(id)
    from forms import LevelThreeEditForm
    form = LevelThreeEditForm()
    if not form.validate_on_submit():
        flash(_t('Please correct the errors.', 'يرجى تصحيح الأخطاء.'), 'danger')
        return redirect(url_for('coa.level_three_list'))
    drawers = form.drawers.data.strip()
    if LevelThree.query.filter(LevelThree.level_two_id == row.level_two_id,
                               db.func.lower(LevelThree.drawers) == drawers.lower(),
                               LevelThree.id != row.id).first():
        flash(_t('Duplicate drawers under the same parent.', 'اسم مكرر ضمن نفس الأب.'), 'danger')
        return redirect(url_for('coa.level_three_list'))
    row.drawers, row.code_length, row.description = drawers, 5, 'Heading Account'
    row.drawers_ar = (form.drawers_ar.data or '').strip() or None
    row.description_ar = (form.description_ar.data or '').strip() or None
    row.status = _status_from(request)
    db.session.commit()
    flash(_t('Level 3 account updated.', 'تم تحديث الحساب.'), 'success')
    return redirect(url_for('coa.level_three_list'))


@coa_bp.route('/level-three/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def level_three_delete(id):
    row = LevelThree.query.get_or_404(id)
    if row.level_fours:
        flash(_t('Cannot delete: this account has Level 4 children.',
                 'لا يمكن الحذف: يحتوي على حسابات فرعية.'), 'danger')
        return redirect(url_for('coa.level_three_list'))
    db.session.delete(row); db.session.commit()
    flash(_t('Level 3 account deleted.', 'تم حذف الحساب.'), 'success')
    return redirect(url_for('coa.level_three_list'))


@coa_bp.route('/level-three/data')
@login_required
def level_three_data():
    return jsonify([r.to_dict() for r in coa_order(LevelThree.query, LevelThree.code).all()])


# ─── LEVEL FOUR ──────────────────────────────────────────────────
@coa_bp.route('/level-four')
@login_required
def level_four_list():
    pg, q_text, parent, sort, status = _paginate_filter_sort(LevelFour, LevelFour.level_three_code, request)
    parents = coa_order(LevelThree.query, LevelThree.code).all()
    level_ones = coa_order(LevelOne.query, LevelOne.code).all()
    return render_template('coa/level_four.html', pg=pg, rows=pg.items,
                           parents=parents, level_ones=level_ones,
                           q=q_text, parent=parent, sort=sort, status=status)


@coa_bp.route('/level-four/next-code')
@login_required
def level_four_next_code():
    p = LevelThree.query.get(request.args.get('level_three_id', type=int))
    return jsonify({'code': next_level_four_code(p) if p else ''})


@coa_bp.route('/level-four/add', methods=['GET', 'POST'])
@login_required
@admin_required
def level_four_add():
    if request.method == 'GET':
        return _reject_get('coa.level_four_list')
    from forms import LevelFourForm
    form = LevelFourForm()
    form.level_three_id.choices = [(r.id, f'{r.code} — {r.drawers}') for r in coa_order(LevelThree.query, LevelThree.code)]
    if not form.validate_on_submit():
        flash(_t('Please correct the errors.', 'يرجى تصحيح الأخطاء.'), 'danger')
        return redirect(url_for('coa.level_four_list'))
    parent = LevelThree.query.get(form.level_three_id.data)
    if not parent:
        flash(_t('Parent not found.', 'الحساب الأب غير موجود.'), 'danger')
        return redirect(url_for('coa.level_four_list'))
    drawers = form.drawers.data.strip()
    if LevelFour.query.filter(LevelFour.level_three_id == parent.id,
                              db.func.lower(LevelFour.drawers) == drawers.lower()).first():
        flash(_t(f'"{drawers}" already exists under {parent.code}.',
                 f'"{drawers}" موجود بالفعل ضمن {parent.code}.'), 'danger')
        return redirect(url_for('coa.level_four_list'))
    db.session.add(LevelFour(code_length=8, level_three_id=parent.id, level_three_code=parent.code,
                             code=next_level_four_code(parent), drawers=drawers,
                             drawers_ar=(form.drawers_ar.data or '').strip() or None,
                             description='Heading Account',
                             description_ar=(form.description_ar.data or '').strip() or None,
                             status=_status_from(request)))
    db.session.commit()
    flash(_t('Level 4 account created.', 'تم إنشاء حساب المستوى الرابع.'), 'success')
    return redirect(url_for('coa.level_four_list'))


@coa_bp.route('/level-four/<int:id>/edit', methods=['POST'])
@login_required
@admin_required
def level_four_edit(id):
    row = LevelFour.query.get_or_404(id)
    from forms import LevelFourEditForm
    form = LevelFourEditForm()
    if not form.validate_on_submit():
        flash(_t('Please correct the errors.', 'يرجى تصحيح الأخطاء.'), 'danger')
        return redirect(url_for('coa.level_four_list'))
    drawers = form.drawers.data.strip()
    if LevelFour.query.filter(LevelFour.level_three_id == row.level_three_id,
                              db.func.lower(LevelFour.drawers) == drawers.lower(),
                              LevelFour.id != row.id).first():
        flash(_t('Duplicate drawers under the same parent.', 'اسم مكرر ضمن نفس الأب.'), 'danger')
        return redirect(url_for('coa.level_four_list'))
    row.drawers, row.code_length, row.description = drawers, 8, 'Heading Account'
    row.drawers_ar = (form.drawers_ar.data or '').strip() or None
    row.description_ar = (form.description_ar.data or '').strip() or None
    row.status = _status_from(request)
    db.session.commit()
    flash(_t('Level 4 account updated.', 'تم تحديث الحساب.'), 'success')
    return redirect(url_for('coa.level_four_list'))


@coa_bp.route('/level-four/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def level_four_delete(id):
    row = LevelFour.query.get_or_404(id)
    if row.level_fives:
        flash(_t('Cannot delete: this account has Level 5 children.',
                 'لا يمكن الحذف: يحتوي على حسابات فرعية.'), 'danger')
        return redirect(url_for('coa.level_four_list'))
    db.session.delete(row); db.session.commit()
    flash(_t('Level 4 account deleted.', 'تم حذف الحساب.'), 'success')
    return redirect(url_for('coa.level_four_list'))


@coa_bp.route('/level-four/data')
@login_required
def level_four_data():
    return jsonify([r.to_dict() for r in coa_order(LevelFour.query, LevelFour.code).all()])


# ─── LEVEL FIVE ──────────────────────────────────────────────────
@coa_bp.route('/level-five')
@login_required
def level_five_list():
    pg, q_text, parent, sort, status = _paginate_filter_sort(LevelFive, LevelFive.level_four_code, request)
    parents = coa_order(LevelFour.query, LevelFour.code).all()
    level_ones = coa_order(LevelOne.query, LevelOne.code).all()
    return render_template('coa/level_five.html', pg=pg, rows=pg.items,
                           parents=parents, level_ones=level_ones,
                           q=q_text, parent=parent, sort=sort, status=status)


@coa_bp.route('/level-five/next-code')
@login_required
def level_five_next_code():
    p = LevelFour.query.get(request.args.get('level_four_id', type=int))
    return jsonify({'code': next_level_five_code(p) if p else ''})


@coa_bp.route('/level-five/add', methods=['GET', 'POST'])
@login_required
@admin_required
def level_five_add():
    if request.method == 'GET':
        return _reject_get('coa.level_five_list')
    from forms import LevelFiveForm
    form = LevelFiveForm()
    form.level_four_id.choices = [(r.id, f'{r.code} — {r.drawers}') for r in coa_order(LevelFour.query, LevelFour.code)]
    if not form.validate_on_submit():
        flash(_t('Please correct the errors.', 'يرجى تصحيح الأخطاء.'), 'danger')
        return redirect(url_for('coa.level_five_list'))
    parent = LevelFour.query.get(form.level_four_id.data)
    if not parent:
        flash(_t('Parent not found.', 'الحساب الأب غير موجود.'), 'danger')
        return redirect(url_for('coa.level_five_list'))
    drawers = form.drawers.data.strip()
    if LevelFive.query.filter(LevelFive.level_four_id == parent.id,
                              db.func.lower(LevelFive.drawers) == drawers.lower()).first():
        flash(_t(f'"{drawers}" already exists under {parent.code}.',
                 f'"{drawers}" موجود بالفعل ضمن {parent.code}.'), 'danger')
        return redirect(url_for('coa.level_five_list'))
    db.session.add(LevelFive(code_length=12, level_four_id=parent.id, level_four_code=parent.code,
                             code=next_level_five_code(parent), drawers=drawers,
                             drawers_ar=(form.drawers_ar.data or '').strip() or None,
                             description='Transactional Account',
                             description_ar=(form.description_ar.data or '').strip() or None,
                             control_account=('Yes' if (form.control_account.data or 'No')=='Yes' else 'No'),
                             status=_status_from(request)))
    db.session.commit()
    flash(_t('Level 5 account created.', 'تم إنشاء حساب المستوى الخامس.'), 'success')
    return redirect(url_for('coa.level_five_list'))


@coa_bp.route('/level-five/<int:id>/edit', methods=['POST'])
@login_required
@admin_required
def level_five_edit(id):
    row = LevelFive.query.get_or_404(id)
    from forms import LevelFiveEditForm
    form = LevelFiveEditForm()
    if not form.validate_on_submit():
        flash(_t('Please correct the errors.', 'يرجى تصحيح الأخطاء.'), 'danger')
        return redirect(url_for('coa.level_five_list'))
    drawers = form.drawers.data.strip()
    if LevelFive.query.filter(LevelFive.level_four_id == row.level_four_id,
                              db.func.lower(LevelFive.drawers) == drawers.lower(),
                              LevelFive.id != row.id).first():
        flash(_t('Duplicate drawers under the same parent.', 'اسم مكرر ضمن نفس الأب.'), 'danger')
        return redirect(url_for('coa.level_five_list'))
    row.drawers, row.code_length, row.description = drawers, 12, 'Transactional Account'
    row.drawers_ar = (form.drawers_ar.data or '').strip() or None
    row.description_ar = (form.description_ar.data or '').strip() or None
    row.control_account = 'Yes' if (form.control_account.data or 'No')=='Yes' else 'No'
    row.status = _status_from(request)
    db.session.commit()
    flash(_t('Level 5 account updated.', 'تم تحديث الحساب.'), 'success')
    return redirect(url_for('coa.level_five_list'))


@coa_bp.route('/level-five/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def level_five_delete(id):
    row = LevelFive.query.get_or_404(id)
    db.session.delete(row); db.session.commit()
    flash(_t('Level 5 account deleted.', 'تم حذف الحساب.'), 'success')
    return redirect(url_for('coa.level_five_list'))


@coa_bp.route('/level-five/data')
@login_required
def level_five_data():
    return jsonify([r.to_dict() for r in coa_order(LevelFive.query, LevelFive.code).all()])


# ═════════════════════════════════════════════════════════════════
#  CASCADING DROPDOWN ENDPOINTS
#  Each returns ONLY the children of the supplied parent, so a form
#  can never offer an invalid hierarchy combination.
#  Response: [{'id':1,'code':'A1','drawers':'...','label':'A1 - ...'}]
# ═════════════════════════════════════════════════════════════════

def _children_payload(rows):
    """Serialise child rows into the `code - Drawers` display format."""
    return jsonify([{
        'id': r.id,
        'code': r.code,
        'drawers': r.drawers,
        'label': f'{r.code} - {r.drawers}',
        'status': r.status or 'active',
    } for r in rows])


@coa_bp.route('/children/level-two')
@login_required
def children_level_two():
    """Level 2 records belonging to the given Level 1."""
    pid = request.args.get('level_one_id', type=int)
    if not pid:
        return jsonify([])
    rows = coa_order(
        LevelTwo.query.filter(LevelTwo.level_one_id == pid, LevelTwo.status == 'active'),
        LevelTwo.code).all()
    return _children_payload(rows)


@coa_bp.route('/children/level-three')
@login_required
def children_level_three():
    """Level 3 records belonging to the given Level 2."""
    pid = request.args.get('level_two_id', type=int)
    if not pid:
        return jsonify([])
    rows = coa_order(
        LevelThree.query.filter(LevelThree.level_two_id == pid, LevelThree.status == 'active'),
        LevelThree.code).all()
    return _children_payload(rows)


@coa_bp.route('/children/level-four')
@login_required
def children_level_four():
    """Level 4 records belonging to the given Level 3."""
    pid = request.args.get('level_three_id', type=int)
    if not pid:
        return jsonify([])
    rows = coa_order(
        LevelFour.query.filter(LevelFour.level_three_id == pid, LevelFour.status == 'active'),
        LevelFour.code).all()
    return _children_payload(rows)


@coa_bp.route('/ancestors/<level>/<int:id>')
@login_required
def ancestors(level, id):
    """Return the full ancestor chain of a record, for EDIT mode.

    The form uses this to repopulate every dropdown in the right order.
    e.g. /coa/ancestors/level-five/12 ->
         {'level_one_id':1,'level_two_id':3,'level_three_id':9,'level_four_id':20}
    """
    out = {}
    if level == 'level-two':
        r = LevelTwo.query.get_or_404(id)
        out = {'level_one_id': r.level_one_id}
    elif level == 'level-three':
        r = LevelThree.query.get_or_404(id)
        out = {'level_two_id': r.level_two_id,
               'level_one_id': r.level_two.level_one_id if r.level_two else None}
    elif level == 'level-four':
        r = LevelFour.query.get_or_404(id)
        l3 = r.level_three
        out = {'level_three_id': r.level_three_id,
               'level_two_id': l3.level_two_id if l3 else None,
               'level_one_id': l3.level_two.level_one_id if (l3 and l3.level_two) else None}
    elif level == 'level-five':
        r = LevelFive.query.get_or_404(id)
        l4 = r.level_four
        l3 = l4.level_three if l4 else None
        l2 = l3.level_two if l3 else None
        out = {'level_four_id': r.level_four_id,
               'level_three_id': l4.level_three_id if l4 else None,
               'level_two_id': l3.level_two_id if l3 else None,
               'level_one_id': l2.level_one_id if l2 else None}
    else:
        return jsonify({'error': 'unknown level'}), 404
    return jsonify(out)


# ═════════════════════════════════════════════════════════════════
#  LIVE VALIDATION ENDPOINTS
#  Power the red ✗ / green ✓ indicators in the forms. They mirror the
#  server-side rules exactly, so the UI can never promise something the
#  POST would later reject.
# ═════════════════════════════════════════════════════════════════

@coa_bp.route('/validate/level-one-code')
@login_required
def validate_level_one_code():
    """A Level 1 code must be a single letter and must not already exist."""
    code = (request.args.get('code') or '').strip().upper()
    exclude = request.args.get('exclude', type=int)   # the row being edited

    if not code:
        return jsonify({'ok': False, 'msg': 'Code is required'})
    if not re.fullmatch(r'[A-Z]', code):
        return jsonify({'ok': False, 'msg': 'Code must be a single letter (A-Z)'})

    q = LevelOne.query.filter(LevelOne.code == code)
    if exclude:
        q = q.filter(LevelOne.id != exclude)
    if q.first():
        return jsonify({'ok': False, 'msg': f'Code "{code}" already exists'})

    return jsonify({'ok': True, 'msg': 'Code is available'})


# level -> (model, parent-id column, label)
_DRAWER_LEVELS = {
    'level-one':   (LevelOne,   None),
    'level-two':   (LevelTwo,   'level_one_id'),
    'level-three': (LevelThree, 'level_two_id'),
    'level-four':  (LevelFour,  'level_three_id'),
    'level-five':  (LevelFive,  'level_four_id'),
}


@coa_bp.route('/validate/<level>/drawers')
@login_required
def validate_drawers(level):
    """Drawers must be non-empty and unique within the selected parent."""
    entry = _DRAWER_LEVELS.get(level)
    if not entry:
        return jsonify({'ok': False, 'msg': 'Unknown level'}), 404
    model, parent_col = entry

    drawers = (request.args.get('drawers') or '').strip()
    parent_id = request.args.get('parent_id', type=int)
    exclude = request.args.get('exclude', type=int)

    if not drawers:
        return jsonify({'ok': False, 'msg': 'Drawers is required'})

    # Without a parent we cannot check for duplicates yet — the name itself
    # is valid, so report success but say the check is pending.
    if parent_col and not parent_id:
        return jsonify({'ok': True, 'msg': 'Select a parent to check for duplicates',
                        'pending': True})

    q = model.query.filter(db.func.lower(model.drawers) == drawers.lower())
    if parent_col and parent_id:
        q = q.filter(getattr(model, parent_col) == parent_id)
    if exclude:
        q = q.filter(model.id != exclude)

    if q.first():
        where = ' under this parent' if parent_col else ''
        return jsonify({'ok': False, 'msg': f'"{drawers}" already exists{where}'})

    return jsonify({'ok': True, 'msg': 'Looks good'})

# ═════════════════════════════════════════════════════════════════
#  CHART OF ACCOUNTS — combined VIEWS (flat + tree)
#  Additive, read-only. Builds the full account hierarchy in Python by
#  joining level_one..level_five on their id columns, so it works with
#  the real schema (level_one..level_five, `drawers`).
# ═════════════════════════════════════════════════════════════════
def _build_coa_maps():
    l1 = {r.id: r for r in LevelOne.query.all()}
    l2 = {r.id: r for r in LevelTwo.query.all()}
    l3 = {r.id: r for r in LevelThree.query.all()}
    l4 = {r.id: r for r in LevelFour.query.all()}
    l5 = coa_order(LevelFive.query, LevelFive.code).all()
    return l1, l2, l3, l4, l5


@coa_bp.route('/views')
@login_required
def coa_views():
    """Page with two tabs: flat account view and tree view."""
    return render_template('coa/views.html')


@coa_bp.route('/views/flat')
@login_required
def coa_views_flat():
    """Flat rows: account_code + the five drawer names (Level 5 leaves)."""
    l1, l2, l3, l4, l5 = _build_coa_maps()
    out = []
    for r5 in coa_sort_list(list(l5)):
        r4 = l4.get(r5.level_four_id)
        r3 = l3.get(r4.level_three_id) if r4 else None
        r2 = l2.get(r3.level_two_id) if r3 else None
        r1 = l1.get(r2.level_one_id) if r2 else None
        out.append({
            'account_code': r5.code,
            'level1_drawer': r1.drawers if r1 else '',
            'level2_drawer': r2.drawers if r2 else '',
            'level3_drawer': r3.drawers if r3 else '',
            'level4_drawer': r4.drawers if r4 else '',
            'level5_drawer': r5.drawers,
            'level1_drawer_ar': (r1.drawers_ar or '') if r1 else '',
            'level2_drawer_ar': (r2.drawers_ar or '') if r2 else '',
            'level3_drawer_ar': (r3.drawers_ar or '') if r3 else '',
            'level4_drawer_ar': (r4.drawers_ar or '') if r4 else '',
            'level5_drawer_ar': r5.drawers_ar or '',
            'description': r5.description or '',
            'description_ar': r5.description_ar or '',
            'control_account': r5.control_account or 'No',
        })
    return jsonify(out)


@coa_bp.route('/views/tree')
@login_required
def coa_views_tree():
    """Nested tree: L1 -> L2 -> L3 -> L4 -> L5 (leaves show code + drawer)."""
    l1, l2, l3, l4, l5 = _build_coa_maps()

    # index children by parent id
    from collections import defaultdict
    c2 = defaultdict(list)   # level_one_id -> [L2]
    c3 = defaultdict(list)   # level_two_id -> [L3]
    c4 = defaultdict(list)   # level_three_id -> [L4]
    c5 = defaultdict(list)   # level_four_id -> [L5]
    for r in l2.values():
        c2[r.level_one_id].append(r)
    for r in l3.values():
        c3[r.level_two_id].append(r)
    for r in l4.values():
        c4[r.level_three_id].append(r)
    for r in l5:
        c5[r.level_four_id].append(r)

    def node(code, name, children, name_ar='', control=None):
        n = {'code': code, 'name': name, 'name_ar': name_ar or '', 'children': children}
        if control is not None:
            n['control_account'] = control
        return n

    tree = []
    for r1 in coa_sort_list(list(l1.values())):
        n2 = []
        for r2 in coa_sort_list(c2.get(r1.id, [])):
            n3 = []
            for r3 in coa_sort_list(c3.get(r2.id, [])):
                n4 = []
                for r4 in coa_sort_list(c4.get(r3.id, [])):
                    n5 = [node(r5.code, r5.drawers, [], r5.drawers_ar,
                               r5.control_account or 'No')
                          for r5 in coa_sort_list(c5.get(r4.id, []))]
                    n4.append(node(r4.code, r4.drawers, n5, r4.drawers_ar))
                n3.append(node(r3.code, r3.drawers, n4, r3.drawers_ar))
            n2.append(node(r2.code, r2.drawers, n3, r2.drawers_ar))
        tree.append(node(r1.code, r1.drawers, n2, r1.drawers_ar))
    return jsonify(tree)

# ══════════════════════════════════════════════════════════════════
#  CHART OF ACCOUNTS — IMPORT / EXPORT  (Levels 1–5)
# ══════════════════════════════════════════════════════════════════
import io as _io
import csv as _csv

# Column layout per level. Parent-code column lets import re-link children.
# (header, model attribute)
_COA_LEVELS = {
    1: {'model': LevelOne,   'sheet': 'Level 1', 'parent': None,
        'cols': [('Code', 'code'), ('Drawers', 'drawers'),
                 ('Drawers (AR)', 'drawers_ar'), ('Description', 'description'),
                 ('Description (AR)', 'description_ar'), ('Status', 'status')]},
    2: {'model': LevelTwo,   'sheet': 'Level 2', 'parent': ('Parent Code (L1)', 'level_one_code', LevelOne, 'level_one_id'),
        'cols': [('Parent Code (L1)', 'level_one_code'), ('Code', 'code'), ('Drawers', 'drawers'),
                 ('Drawers (AR)', 'drawers_ar'), ('Description', 'description'),
                 ('Description (AR)', 'description_ar'), ('Status', 'status')]},
    3: {'model': LevelThree, 'sheet': 'Level 3', 'parent': ('Parent Code (L2)', 'level_two_code', LevelTwo, 'level_two_id'),
        'cols': [('Parent Code (L2)', 'level_two_code'), ('Code', 'code'), ('Drawers', 'drawers'),
                 ('Drawers (AR)', 'drawers_ar'), ('Description', 'description'),
                 ('Description (AR)', 'description_ar'), ('Status', 'status')]},
    4: {'model': LevelFour,  'sheet': 'Level 4', 'parent': ('Parent Code (L3)', 'level_three_code', LevelThree, 'level_three_id'),
        'cols': [('Parent Code (L3)', 'level_three_code'), ('Code', 'code'), ('Drawers', 'drawers'),
                 ('Drawers (AR)', 'drawers_ar'), ('Description', 'description'),
                 ('Description (AR)', 'description_ar'), ('Status', 'status')]},
    5: {'model': LevelFive,  'sheet': 'Level 5', 'parent': ('Parent Code (L4)', 'level_four_code', LevelFour, 'level_four_id'),
        'cols': [('Parent Code (L4)', 'level_four_code'), ('Code', 'code'), ('Drawers', 'drawers'),
                 ('Drawers (AR)', 'drawers_ar'), ('Description', 'description'),
                 ('Description (AR)', 'description_ar'), ('Control Account', 'control_account'),
                 ('Status', 'status')]},
}


@coa_bp.route('/export')
@login_required
def coa_export():
    """Export all five levels to a single .xlsx (one sheet per level),
    or to CSV of one level with ?fmt=csv&level=N."""
    fmt = (request.args.get('fmt') or 'xlsx').lower()

    if fmt == 'csv':
        try:
            level = int(request.args.get('level', 1))
        except (TypeError, ValueError):
            level = 1
        spec = _COA_LEVELS.get(level)
        if not spec:
            return jsonify({'ok': False, 'error': 'Invalid level'}), 400
        rows = coa_order(spec['model'].query, spec['model'].code).all()
        buf = _io.StringIO()
        w = _csv.writer(buf)
        w.writerow([h for h, _ in spec['cols']])
        for r in rows:
            w.writerow([getattr(r, attr, '') or '' for _, attr in spec['cols']])
        data = buf.getvalue().encode('utf-8-sig')   # BOM for Excel/Arabic
        return send_file(_io.BytesIO(data), as_attachment=True,
                         download_name=f'chart_of_accounts_level_{level}.csv',
                         mimetype='text/csv')

    # xlsx: one sheet per level
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    hdr_fill = PatternFill('solid', fgColor='1E3A5F')
    hdr_font = Font(color='FFFFFF', bold=True, size=10)

    for level in (1, 2, 3, 4, 5):
        spec = _COA_LEVELS[level]
        ws = wb.create_sheet(title=spec['sheet'])
        for i, (h, _) in enumerate(spec['cols'], 1):
            c = ws.cell(row=1, column=i, value=h)
            c.fill = hdr_fill
            c.font = hdr_font
            ws.column_dimensions[get_column_letter(i)].width = max(14, len(h) + 3)
        ws.freeze_panes = 'A2'
        rows = coa_order(spec['model'].query, spec['model'].code).all()
        for r_i, row in enumerate(rows, 2):
            for c_i, (_, attr) in enumerate(spec['cols'], 1):
                ws.cell(row=r_i, column=c_i, value=getattr(row, attr, '') or '')

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name='chart_of_accounts.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def _coa_import_rows(level, rows):
    """Insert new accounts for one level. rows = list of dicts keyed by header.
    Add-new-only: existing codes are skipped. Returns (added, skipped, errors)."""
    spec = _COA_LEVELS[level]
    model = spec['model']
    added = skipped = 0
    errors = []

    # Cache existing codes for this level.
    existing = {r.code for r in model.query.with_entities(model.code).all()}

    parent = spec['parent']
    parent_map = {}
    if parent:
        _, _, pmodel, _ = parent
        parent_map = {p.code: p.id for p in pmodel.query.with_entities(pmodel.code, pmodel.id).all()}

    for idx, row in enumerate(rows, 2):
        code = (str(row.get('Code', '')) or '').strip()
        if not code:
            continue   # skip blank lines silently
        if code in existing:
            skipped += 1
            continue

        obj = model()
        obj.code = code
        obj.drawers = (str(row.get('Drawers', '')) or '').strip()
        obj.drawers_ar = (str(row.get('Drawers (AR)', '')) or '').strip() or None
        obj.description = (str(row.get('Description', '')) or '').strip() or _default_desc(level)
        obj.description_ar = (str(row.get('Description (AR)', '')) or '').strip() or None
        obj.status = (str(row.get('Status', '')) or 'active').strip().lower()
        if obj.status not in ('active', 'inactive'):
            obj.status = 'active'
        obj.code_length = {1: 1, 2: 2, 3: 5, 4: 8, 5: 12}[level]

        if level == 5:
            ca = (str(row.get('Control Account', '')) or 'No').strip()
            obj.control_account = 'Yes' if ca.lower() in ('yes', 'y', '1', 'true') else 'No'

        if parent:
            phdr, pcode_attr, pmodel, pid_attr = parent
            pcode = (str(row.get(phdr, '')) or '').strip()
            if not pcode or pcode not in parent_map:
                errors.append(f'Row {idx} ({code}): parent "{pcode}" not found; skipped.')
                continue
            setattr(obj, pcode_attr, pcode)
            setattr(obj, pid_attr, parent_map[pcode])

        db.session.add(obj)
        existing.add(code)
        added += 1

    return added, skipped, errors


def _default_desc(level):
    return 'Elements of Financial Statements' if level == 1 else (
        'Transactional Account' if level == 5 else 'Heading Account')


@coa_bp.route('/import', methods=['POST'])
@login_required
def coa_import():
    """Import accounts from an uploaded .xlsx (all 5 sheets) or .csv (one level).
    Add-new-only: existing codes are skipped. Import runs parent-first (L1→L5)."""
    f = request.files.get('file')
    if not f or not f.filename:
        return jsonify({'ok': False, 'error': 'No file uploaded.'}), 400

    name = f.filename.lower()
    total_added = total_skipped = 0
    all_errors = []

    try:
        if name.endswith('.csv'):
            try:
                level = int(request.form.get('level', 0))
            except (TypeError, ValueError):
                level = 0
            if level not in _COA_LEVELS:
                return jsonify({'ok': False, 'error': 'For CSV, choose which level it belongs to.'}), 400
            text = f.read().decode('utf-8-sig', errors='replace')
            reader = _csv.DictReader(_io.StringIO(text))
            rows = list(reader)
            a, s, e = _coa_import_rows(level, rows)
            total_added += a; total_skipped += s; all_errors += e

        elif name.endswith('.xlsx') or name.endswith('.xlsm'):
            from openpyxl import load_workbook
            wb = load_workbook(f, data_only=True, read_only=True)
            # Process parent-first so child parent-codes resolve.
            for level in (1, 2, 3, 4, 5):
                sheet = _COA_LEVELS[level]['sheet']
                if sheet not in wb.sheetnames:
                    continue
                ws = wb[sheet]
                it = ws.iter_rows(values_only=True)
                try:
                    headers = [str(h).strip() if h is not None else '' for h in next(it)]
                except StopIteration:
                    continue
                rows = []
                for raw in it:
                    rows.append({headers[i]: raw[i] for i in range(min(len(headers), len(raw)))})
                a, s, e = _coa_import_rows(level, rows)
                total_added += a; total_skipped += s; all_errors += e
        else:
            return jsonify({'ok': False, 'error': 'Unsupported file type. Use .xlsx or .csv'}), 400

        db.session.commit()
        return jsonify({'ok': True, 'added': total_added, 'skipped': total_skipped,
                        'errors': all_errors[:50]})
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 500